"""Shared plumbing for server-side report pages (PRD §6 / §10).

Contract with the frontend (planing/frontend_tasks.md):
- deferred prop `report` = { rows, total, summary, options, conn_error }
- prop `filters` echoes the requested params so the form can be restored
- GET <page-url>/export?<same query> answers with an XLSX file.
"""
import datetime as dt
import os
import re
from decimal import Decimal

from django.http import HttpResponse

# Karakter kontrol yg ditolak XML 1.0 -> openpyxl melempar IllegalCharacterError
# (bukan pyodbc.Error, jadi lolos dari `except pyodbc.Error` di view = 500).
# Teks legacy MS SQL (keterangan/nama/alamat) sering bawa byte ini. Cermin dari
# openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE; didefinisikan lokal agar openpyxl tak
# ikut ter-import di tiap request (dipakai di _clean, hot loop per-sel).
_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _env_int(name, default):
    try:
        return int(os.environ[name])
    except (KeyError, ValueError):
        return default


# Semua batas bisa ditune lewat env (mis. REPORT_MAX_PER_PAGE) tanpa ubah kode.
DEFAULT_PER_PAGE = _env_int("REPORT_DEFAULT_PER_PAGE", 100)
MAX_PER_PAGE = _env_int("REPORT_MAX_PER_PAGE", 1000)     # plafon page-size di layar (dulu 200)
MAX_RANGE_DAYS = _env_int("REPORT_MAX_RANGE_DAYS", 92)   # ~3 bln; hanya clamp jalur INTERAKTIF, export dilepas
EXPORT_CAP = _env_int("REPORT_EXPORT_CAP", 100_000)      # cap XLSX (buffer RAM); CSV streaming TIDAK dibatasi ini
RECENT_LIMIT = _env_int("REPORT_RECENT_LIMIT", 100)      # rows shown on first load, before any filter is set

RANGE_WARNING = (
    "Rentang tanggal dibatasi maksimal {d} hari — tanggal mulai disesuaikan otomatis."
)


def _parse_date(s):
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d") if s else None
    except (ValueError, TypeError):
        return None


def month_start(today=None) -> dt.datetime:
    today = today or dt.date.today()
    return dt.datetime(today.year, today.month, 1)


def parse_report_params(
    request, sorts, default_sort, max_range_days=MAX_RANGE_DAYS,
    enable_recent=False, recent_sort=None, default_from_days=None,
    default_sort_dir="desc",
):
    """Read the standard report params from request.GET into a dict `f`.

    `sorts` is the whitelist: {sort_param: output_column_alias}. Anything not in
    it falls back to `default_sort` — this is the SQL-injection guard, since the
    ORDER BY clause is built from the alias, never from raw user input.

    `default_from_days`: bila di-set, `date_from` bawaan (saat request tak
    membawa tanggal) jadi `date_to - N hari` alih-alih awal bulan. Laporan yang
    menilai RIWAYAT pelanggan — bukan penjualan satu periode — butuh jendela
    panjang sejak muat pertama; awal-bulan akan mengklasifikasi semua orang
    sebagai baru. `None` = perilaku month_start seperti laporan lainnya.

    `enable_recent`: when True and the request carries no query params at all
    (first page load, before any filter/sort/page interaction), `f["recent"]`
    is set and the report shows the latest `RECENT_LIMIT` rows (see
    `run_recent()`) instead of the default month-to-date page. `recent_sort`
    (a key in `sorts`) picks the column ordered DESC in that mode — reports
    with no natural date column to sort by (aggregates without a raw
    `tanggal`) should leave it unset and fall back to their normal sort.
    """
    g = request.GET
    recent = enable_recent and not any(v for v in g.values())

    today = dt.date.today()
    date_mode = "exact" if g.get("date_mode") == "exact" else "range"
    if date_mode == "exact":
        single = _parse_date(g.get("date")) or dt.datetime(today.year, today.month, today.day)
        date_from = date_to = single
    else:
        date_to = _parse_date(g.get("date_to")) or dt.datetime(today.year, today.month, today.day)
        date_from = _parse_date(g.get("date_from"))
        if date_from is None:
            date_from = (date_to - dt.timedelta(days=default_from_days)
                         if default_from_days else month_start(today))

    # max_range_days=None -> tanpa clamp (jalur export/CSV: akses rentang berapapun).
    warning = None
    if max_range_days is not None and (date_to - date_from).days > max_range_days:
        date_from = date_to - dt.timedelta(days=max_range_days)
        warning = RANGE_WARNING.format(d=max_range_days)

    sort = g.get("sort") or default_sort
    if sort not in sorts:
        sort = default_sort
    # `default_sort_dir` ada karena "terbesar dulu" tidak selalu berarti "yang
    # paling perlu dilihat dulu". Pada kolom yang isinya peringkat urgensi
    # (Klasifikasi Pelanggan: 1=Hilang .. 5=Aktif) desc justru membuang yang
    # paling mendesak ke halaman terakhir.
    sort_dir = "asc" if (g.get("sort_dir") or default_sort_dir).lower() == "asc" else "desc"

    if recent:
        page, per_page = 1, RECENT_LIMIT
        order_by = f"q.{sorts[recent_sort]} DESC" if recent_sort in sorts else f"q.{sorts[sort]} {sort_dir.upper()}"
    else:
        try:
            page = max(1, int(g.get("page") or 1))
        except ValueError:
            page = 1
        try:
            per_page = min(MAX_PER_PAGE, max(10, int(g.get("per_page") or DEFAULT_PER_PAGE)))
        except ValueError:
            per_page = DEFAULT_PER_PAGE
        order_by = f"q.{sorts[sort]} {sort_dir.upper()}"

    return {
        "recent": recent,
        "skip_date_predicate": recent,
        "date_mode": date_mode,
        "date_from": date_from,
        "date_to": date_to.replace(hour=23, minute=59, second=59),
        "date_from_s": date_from.strftime("%Y-%m-%d"),
        "date_to_s": date_to.strftime("%Y-%m-%d"),
        "search": (g.get("search") or "").strip(),
        "sort": sort,
        # Whitelist yang sama dikirim ke layar: dulu tiap halaman Vue menandai
        # sendiri kolom mana yang `sortable`, dan daftar itu menyimpang dari
        # `sorts` di sini — header yang bisa diklik tapi tak mengubah apa pun
        # (sort tak dikenal diam-diam jatuh ke default), dan kolom yang
        # sebenarnya bisa diurut tapi tak pernah ditawarkan.
        "sort_keys": list(sorts),
        "sort_dir": sort_dir,
        "order_by": order_by,
        "page": page,
        "per_page": per_page,
        "warning": warning,
        "export": False,
    }


def dictify(cur) -> list[dict]:
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _clean(v):
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return _ILLEGAL_XML.sub("", v).strip()
    return v


def clean_rows(rows: list[dict]) -> list[dict]:
    """Decimal → float, datetime → string, char() padding stripped."""
    return [{k: _clean(v) for k, v in r.items()} for r in rows]


def one_row(cur) -> dict:
    """Baris pertama hasil query ringkasan, atau {} bila kosong.

    Ringkasan agregat selalu mengembalikan satu baris, jadi `[0]` langsung
    memang tak pernah meledak hari ini — tapi itu IndexError laten yang jadi
    500, dan kelas bug yang sama pernah kejadian (lihat _ILLEGAL_XML di atas)."""
    return (clean_rows(dictify(cur)) or [{}])[0]


def run_paged(cur, inner_sql, params, f):
    """COUNT + OFFSET/FETCH over `inner_sql` (a full SELECT without ORDER BY).

    ORDER BY references the subquery's output aliases (f["order_by"] was built
    from the whitelist in parse_report_params). Returns (rows, total); every row
    gets a synthetic `_rid` (global row number) usable as a stable row key.
    """
    cur.execute(f"SELECT COUNT(*) FROM ({inner_sql}) AS q", params)
    total = int(cur.fetchone()[0] or 0)
    offset = (f["page"] - 1) * f["per_page"]
    cur.execute(
        f"SELECT * FROM ({inner_sql}) AS q ORDER BY {f['order_by']} "
        "OFFSET ? ROWS FETCH NEXT ? ROWS ONLY",
        list(params) + [offset, f["per_page"]],
    )
    rows = clean_rows(dictify(cur))
    for i, r in enumerate(rows):
        r["_rid"] = offset + i + 1
    return rows, total


def run_all(cur, inner_sql, params, f):
    """Export path: the same query without pagination, capped at EXPORT_CAP."""
    cur.execute(
        f"SELECT TOP {EXPORT_CAP} * FROM ({inner_sql}) AS q ORDER BY {f['order_by']}",
        params,
    )
    rows = clean_rows(dictify(cur))
    for i, r in enumerate(rows):
        r["_rid"] = i + 1
    return rows


def run_recent(cur, inner_sql, params, f):
    """First-load path (f["recent"]): TOP `per_page` rows, no COUNT(*) — a
    virgin page load has no filter to narrow the scan, so counting the whole
    table just to throw the number away isn't worth the round-trip. Returns
    (rows, total, capped_sql) — `capped_sql` lets the caller compute a summary
    aggregate over the same capped set rather than the full history.
    """
    # Alias `q` (not q0) so f["order_by"] — built with a `q.` prefix in
    # parse_report_params, matching run_paged/run_all — binds here too.
    capped_sql = f"SELECT TOP {f['per_page']} * FROM ({inner_sql}) AS q ORDER BY {f['order_by']}"
    cur.execute(capped_sql, params)
    rows = clean_rows(dictify(cur))
    for i, r in enumerate(rows):
        r["_rid"] = i + 1
    return rows, len(rows), capped_sql


_FILTER_TYPES = ("text", "number_range", "date", "category")


def parse_column_filters(request, filters_whitelist):
    """Parse per-column advanced filters from request.GET against a whitelist.

    `filters_whitelist`: {name: (sql_alias, type)} — `sql_alias` is always a
    Python literal from this dict, never taken from the request, so it's safe
    to interpolate into SQL (same guard shape as the `sorts` whitelist). Only
    filters with an actual value present are returned.
    """
    g = request.GET
    parsed = {}
    for name, (alias, kind) in filters_whitelist.items():
        key = f"f_{name}"
        if kind == "text":
            v = (g.get(key) or "").strip()
            if v:
                parsed[name] = {"alias": alias, "kind": kind, "value": v}
        elif kind == "number_range":
            try:
                vmin = float(g[f"{key}_min"]) if g.get(f"{key}_min") else None
            except ValueError:
                vmin = None
            try:
                vmax = float(g[f"{key}_max"]) if g.get(f"{key}_max") else None
            except ValueError:
                vmax = None
            if vmin is not None or vmax is not None:
                parsed[name] = {"alias": alias, "kind": kind, "min": vmin, "max": vmax}
        elif kind == "date":
            if (g.get(f"{key}_mode") or "range") == "exact":
                d = _parse_date(g.get(f"{key}_date"))
                if d:
                    parsed[name] = {"alias": alias, "kind": kind, "mode": "exact", "date": d}
            else:
                d_from, d_to = _parse_date(g.get(f"{key}_from")), _parse_date(g.get(f"{key}_to"))
                if d_from or d_to:
                    parsed[name] = {"alias": alias, "kind": kind, "mode": "range", "from": d_from, "to": d_to}
        elif kind == "category":
            values = [v for v in g.getlist(key) if v]
            if values:
                parsed[name] = {"alias": alias, "kind": kind, "values": values}
    return parsed


def apply_column_filters(inner_sql, params, f):
    """AND-combine the parsed advanced filters (f["filters"]) onto `inner_sql`.

    Wraps in an outer SELECT so filters can reference computed/aliased output
    columns (e.g. `subtotal`), which SQL Server can't filter on on the same
    SELECT level they're defined in. Returns `inner_sql` unchanged when there's
    nothing to filter, to avoid an extra subquery layer in the common case.
    """
    parsed = f.get("filters") or {}
    if not parsed:
        return inner_sql, list(params)

    clauses, extra_params = [], []
    for spec in parsed.values():
        alias, kind = spec["alias"], spec["kind"]
        if kind == "text":
            clauses.append(f"q.{alias} LIKE ?")
            extra_params.append(f"%{spec['value']}%")
        elif kind == "number_range":
            if spec["min"] is not None:
                clauses.append(f"q.{alias} >= ?")
                extra_params.append(spec["min"])
            if spec["max"] is not None:
                clauses.append(f"q.{alias} <= ?")
                extra_params.append(spec["max"])
        elif kind == "date":
            if spec["mode"] == "exact":
                clauses.append(f"CAST(q.{alias} AS DATE) = ?")
                extra_params.append(spec["date"])
            else:
                if spec["from"]:
                    clauses.append(f"q.{alias} >= ?")
                    extra_params.append(spec["from"])
                if spec["to"]:
                    clauses.append(f"q.{alias} <= ?")
                    extra_params.append(spec["to"].replace(hour=23, minute=59, second=59))
        elif kind == "category":
            clauses.append(f"q.{alias} IN ({','.join('?' * len(spec['values']))})")
            extra_params.extend(spec["values"])

    wrapped = f"SELECT * FROM ({inner_sql}) AS q WHERE {' AND '.join(clauses)}"
    return wrapped, list(params) + extra_params


def opt(rows, value_key, label_key):
    """Shape master rows into [{value,label}] for SelectSearch dropdowns."""
    out = []
    for r in rows:
        v = r.get(value_key)
        label = r.get(label_key)
        out.append({
            "value": str(v).strip() if v is not None else "",
            "label": str(label).strip() if label is not None else "",
        })
    return out


def xlsx_response(filename, columns, rows):
    """Stream rows as an XLSX download. `columns` = [{key,label}] like the frontend."""
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append([c["label"] for c in columns])
    for r in rows[:EXPORT_CAP]:
        # _clean = buang karakter kontrol ilegal (openpyxl menolaknya) + normalisasi
        # Decimal/tanggal; jalur stream sudah lewat _clean, samakan di sini.
        ws.append([_clean(r.get(c["key"])) for c in columns])
    return _xlsx_download(wb, filename)


def _xlsx_download(wb, filename):
    """Serialisasi workbook jadi response download (stempel waktu di nama file)."""
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M")
    resp["Content-Disposition"] = f'attachment; filename="{filename}-{stamp}.xlsx"'
    wb.save(resp)
    return resp


def _fill_sheet(ws, columns, cur, chunk=2000):
    """Tulis header + seluruh baris `cur` ke `ws`, streaming per `chunk`.

    Kolom dipetakan lewat INDEX dari cur.description, bukan dict per baris —
    itulah yang menjaga memori app datar walau ratusan ribu baris."""
    ws.append([c["label"] for c in columns])
    pos = {name: i for i, name in enumerate(d[0] for d in cur.description)}
    idxs = [pos.get(c["key"]) for c in columns]
    while True:
        batch = cur.fetchmany(chunk)
        if not batch:
            break
        for row in batch:
            ws.append([_clean(row[i]) if i is not None else None for i in idxs])


def xlsx_multi_sheet_response(filename, sheets):
    """XLSX dgn BEBERAPA sheet, masing-masing dari list-of-dict di memori.

    `sheets` = [(title, columns, rows)].

    Berbasis memori, bukan streaming cursor seperti `xlsx_stream_response`, dan
    itu pilihan sadar: satu file multi-sheet hanya masuk akal kalau sheet-nya
    saling berkaitan, dan menjaga kaitan itu (mis. sheet 2 hanya memuat
    pelanggan yang ada di sheet 1) menuntut caller sudah memegang baris sheet 1
    sebagai himpunan. Karena itu pemakainya wajib berpopulasi terbatas —
    Klasifikasi Pelanggan ~4.800 baris. Untuk export ratusan ribu baris tetap
    pakai `xlsx_stream_response` satu sheet.
    """
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    for title, columns, rows in sheets:
        # Excel menolak judul sheet > 31 karakter / berisi : \ / ? * [ ]
        ws = wb.create_sheet(re.sub(r"[:\\/?*\[\]]", "-", title)[:31])
        ws.append([c["label"] for c in columns])
        for r in rows[:EXPORT_CAP]:
            ws.append([_clean(r.get(c["key"])) for c in columns])
    return _xlsx_download(wb, filename)


def xlsx_stream_response(filename, columns, cur, chunk=2000):
    """Bangun XLSX dgn STREAMING cursor (yg sudah di-execute caller) baris per
    baris ke openpyxl write_only — TAK pernah menumpuk seluruh baris jadi
    list-of-dict di RAM (beda `run_all` + `xlsx_response`). Baris diambil sbg
    tuple langsung dari cursor lalu dipetakan ke kolom lewat index. Output XLSX,
    memori app datar walau ratusan ribu baris.
    """
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    _fill_sheet(wb.create_sheet("Data"), columns, cur, chunk=chunk)
    return _xlsx_download(wb, filename)
